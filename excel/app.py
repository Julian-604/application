import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl

class ExcelEditor:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Editor")
        self.file_path = None
        
        self.load_button = ttk.Button(root, text="Load Excel", command=self.load_excel)
        self.load_button.pack(pady=5)
        
        self.tree = ttk.Treeview(root, show='headings')
        self.tree.pack(expand=True, fill='both')
        
        self.add_button = ttk.Button(root, text="Add Entry", command=self.add_entry)
        self.add_button.pack(pady=5)
        
        self.modify_button = ttk.Button(root, text="Modify Entry", command=self.modify_entry)
        self.modify_button.pack(pady=5)
        
        self.delete_button = ttk.Button(root, text="Delete Entry", command=self.delete_entry)
        self.delete_button.pack(pady=5)
    
    def load_excel(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not self.file_path:
            return
        
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.sheet = self.workbook.active
        
        self.tree.delete(*self.tree.get_children())
        self.tree['columns'] = [self.sheet.cell(row=1, column=i).value for i in range(1, self.sheet.max_column + 1)]
        
        for col in self.tree['columns']:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)
        
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            self.tree.insert('', tk.END, values=row)
    
    def add_entry(self):
        if not self.file_path:
            messagebox.showerror("Error", "Load an Excel file first!")
            return
        
        new_data = [tk.simpledialog.askstring("Input", f"Enter {col}:") for col in self.tree['columns']]
        if None not in new_data:
            self.tree.insert('', tk.END, values=new_data)
            self.sheet.append(new_data)
            self.workbook.save(self.file_path)
    
    def modify_entry(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Select an entry to modify")
            return
        
        values = self.tree.item(selected_item, 'values')
        updated_values = [tk.simpledialog.askstring("Modify", f"Modify {col}:", initialvalue=values[i]) for i, col in enumerate(self.tree['columns'])]
        
        if None not in updated_values:
            self.tree.item(selected_item, values=updated_values)
            row_index = self.tree.index(selected_item) + 2
            for col_idx, new_value in enumerate(updated_values, start=1):
                self.sheet.cell(row=row_index, column=col_idx, value=new_value)
            self.workbook.save(self.file_path)
    
    def delete_entry(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Select an entry to delete")
            return
        
        self.tree.delete(selected_item)
        row_index = self.tree.index(selected_item) + 2
        self.sheet.delete_rows(row_index)
        self.workbook.save(self.file_path)
        
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelEditor(root)
    root.mainloop()
